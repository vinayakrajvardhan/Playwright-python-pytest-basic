'''
openpyxl
    pip install openpyxl
'''
import openpyxl
import pytest
from playwright.sync_api import expect,Page

# reading data from excel
login_data=[]

workbook=openpyxl.load_workbook("testdata/data.xlsx")
sheet=workbook.active    # or  worksheet["Sheetname"]

# reading data from excel sheet using for loop
for row in sheet.iter_rows(min_row=2, values_only=True):
    login_data.append(row)

# for row in sheet.iter_rows(min_row=2, values_only=True):
#     email, password, validity = row
#     login_data.append((str(email or ""), str(password or ""), str(validity or "")))
# workbook.close()

@pytest.mark.parametrize("email,password,validity",login_data)
def test_login_data_driven_excel(email,password,validity,page:Page):
    page.goto("https://demowebshop.tricentis.com/login")

    #fill teh login form
    page.locator("#Email").fill(email)   # email id
    page.locator("#Password").fill(password)  #password
    page.locator("input[value='Log in']").click()

    #validation
    if validity=="valid":
        logout_link=page.locator("a[href='/logout']")
        expect(logout_link).to_be_visible(timeout=5000)
    else:
        error_message=page.locator(".validation-summary-errors")
        expect(error_message).to_be_visible(timeout=5000) # checking error message
        expect(page).to_have_url("https://demowebshop.tricentis.com/login") # checking same login page








